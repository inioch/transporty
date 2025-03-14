from pandas import DataFrame
import csv
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Transporty")
        self.root.geometry("400x300")

        self.file_path = None
        self.tables = {}

        self.create_widgets()

    def create_widgets(self):
        tk.Label(self.root, text="Wybierz plik CSV:").pack()
        tk.Button(self.root, text="Kolumny", command=self.select_file).pack()

    def select_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if file_path:
            response = messagebox.askquestion("Wybrany plik", f"Czy na pewno chcesz przekonwertować ten plik?")
            if response == "yes":
                self.convert_file(file_path)
            else:
                print("Nie")
        else:
            messagebox.showinfo("Plik nie wybrany", "Nie wybrano pliku")

    def convert_file(self, file_path):
        selected_headers = ['LP', 'AWB', 'Parts', 'Weight', 'Name', 'Address', 'Town', 'Postcode', 'Number', "Product"]
        selected_columns = [2, 4, 10, 11, 26, 27, 29, 30, 34, 52]

        try:
            tables = {}

            with open(file_path, "r", encoding="utf-8") as file:
                reader = csv.reader(file)
                for row in reader:
                    if not row:
                        continue
                    key = row[0].strip('"')
                    if key not in tables:
                        tables[key] = []
                    tables[key].append(row)

            if "A" in tables:
                df_a = DataFrame(tables["A"])

                if len(selected_columns) != len(selected_headers):
                    messagebox.showerror("Błąd", "Liczba nagłówków nie zgadza się z liczbą kolumn!")
                    return

                df_a_selected = df_a.iloc[:, selected_columns]
                df_a_selected.columns = selected_headers

                # Zapis do Excela i ustawianie szerokości kolumn
                file_name = "A_clean.xlsx"
                wb = Workbook()
                ws = wb.active
                ws.title = "Dane"

                ws.append(selected_headers)  # Dodanie nagłówków

                for row in df_a_selected.itertuples(index=False, name=None):
                    ws.append(row)

                # **Ustawienie szerokości kolumn**
                for col_idx, col_letter in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
                    max_length = max((len(str(cell.value)) for cell in col_letter if cell.value), default=10)
                    ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

                    # **Formatowanie pierwszych dwóch kolumn jako liczby**
                    if col_idx in [1, 2]:  # Kolumny LP i AWB
                        for cell in col_letter:
                            if cell.value is not None:
                                try:
                                    cell.value = int(cell.value)  # Konwersja na liczbę całkowitą
                                    cell.number_format = '0'  # Format bez miejsc po przecinku
                                except ValueError:
                                    pass  # Jeśli nie można przekonwertować, zostawiamy wartość bez zmian

                wb.save(file_name)

                messagebox.showinfo("Zapisano", f"Dane zostały zapisane do pliku {file_name}")
            else:
                messagebox.showinfo("Brak A w pliku CSV", "Brak A w pliku CSV. Wczytano niepoprawny plik")

        except FileNotFoundError:
            messagebox.showerror("Błąd", f"Plik {file_path} nie istnieje.")
        except Exception as e:
            messagebox.showerror("Błąd", f"Wystąpił błąd: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
